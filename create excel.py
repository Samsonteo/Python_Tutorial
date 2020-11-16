from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

newwb = Workbook()
ws = newwb.active
new_sheet = newwb.create_sheet("newSheet")

c_line=0; line_split = []; max_row=0; min_row=0; max=0; min=0;

thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
border = Border(top=double, left=thin, right=thin, bottom=double)
fill = PatternFill("solid", fgColor="DDDDDD")
fill = GradientFill(stop=("000000", "FFFFFF"))
font = Font(b=True, color="FF0000")
al = Alignment(horizontal="center", vertical="center")

file=open("sales-report.txt", "r")
ws.cell(row=1, column=1).value = "SALES REPORT"
ws.cell(row=2, column=1).value = "Name"
ws.cell(row=2, column=2).value = "State"
ws.cell(row=2, column=3).value = "Sales"

for line in file:
	line_split.append(line.split("|"))
	ws.cell(row=c_line+3, column=1).value = line_split[c_line][0]
	ws.cell(row=c_line+3, column=2).value = line_split[c_line][1]
	ws.cell(row=c_line+3, column=3).value = float(line_split[c_line][2])
	ws.cell(row=c_line+3, column=2).alignment = al
	ws.cell(row=c_line+3, column=3).number_format = '#,##0.00'
	if c_line == 0:
		max_row = c_line+3; min_row = c_line+3; 
		max = min = float(line_split[c_line][2]);
	if max < float(line_split[c_line][2]):
		max = float(line_split[c_line][2]); max_row = c_line+3;
	if min > float(line_split[c_line][2]):
		min = float(line_split[c_line][2]); min_row = c_line+3; 
	c_line+=1	
file.close()

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

ws.row_dimensions[2].font = font
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15
ws.column_dimensions["A"].alignment = al

maxFill = PatternFill(start_color='0000FF00', end_color='0000FF00',  fill_type='solid')
minFill = PatternFill(start_color='00FF0000', end_color='00FF0000',  fill_type='solid')
for x in range(1,4):
	ws.cell(row=2,column=x).font = font
	ws.cell(row=2,column=x).alignment = al
	ws.cell(row=max_row, column=x).fill = maxFill
	ws.cell(row=min_row, column=x).fill = minFill
style_range(ws, 'A1:C1', border=border, fill=fill, font=font, alignment=al)
print("\n\n\n************************************************************")
print("********    Importing records to Excel .....     ***********")
print("************************************************************")
newwb.save("sales-report.xlsx")
print("\n\n\nSales-report.xlsx Successfull Created")
